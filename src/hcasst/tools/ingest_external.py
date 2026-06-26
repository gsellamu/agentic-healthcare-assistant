"""Ingest instructor-provided datasets into typed PatientHistory objects.

The course dataset ships patient data in two shapes:
  - records.xlsx           : roster (name, age, gender, contact, free-text summary)
  - sample_report_*.pdf    : SOAP "History and Physical Note" per patient

Both are adapted into the same Patient / EvidenceItem / PatientHistory models the
rest of the system already uses — so RAG, summarization, the grounding guardrail,
and booking all work on real data unchanged. This is the payoff of programming to
the model interface rather than to a single data source.

Patient ids are slugged from the name (e.g. "Anjali Mehra" -> "ext-anjali-mehra").
A SOAP report, when present, is preferred over the xlsx summary for that patient
because it is richer and itemizable.

Note on the messy CCD (sample_patient.pdf, Rebeca Nagle): only the headline
diagnosis/medications are extracted; full labs/vitals/multi-encounter parsing is
a deliberate non-goal (a robust CCD parser is its own project).
"""

from __future__ import annotations

import re
from pathlib import Path

from config.settings import DATASETS_DIR
from hcasst.models import EvidenceItem, EvidenceKind, Patient, PatientHistory


def _slug(name: str) -> str:
    return "ext-" + re.sub(r"[^a-z0-9]+", "-", name.lower()).strip("-")


def _dob_to_iso(dob: str) -> str:
    """MM/DD/YYYY -> YYYY-MM-DD; pass through anything else."""
    m = re.match(r"(\d{1,2})/(\d{1,2})/(\d{4})", dob.strip())
    if m:
        mm, dd, yyyy = m.groups()
        return f"{yyyy}-{int(mm):02d}-{int(dd):02d}"
    return dob.strip()


# ── records.xlsx ──────────────────────────────────────────────────────────
def load_xlsx_patients(path: Path | None = None) -> list[PatientHistory]:
    path = Path(path or DATASETS_DIR / "records.xlsx")
    if not path.exists():
        return []
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    idx = {name: header.index(name) for name in header}

    seen: dict[str, PatientHistory] = {}
    for raw in rows[1:]:
        rec = {h: (raw[idx[h]] if idx[h] < len(raw) else None) for h in header}
        name = str(rec.get("Name") or "").strip()
        if not name:
            continue
        pid = _slug(name)
        if pid in seen:                       # xlsx has duplicate rows; keep first
            continue
        patient = Patient(
            id=pid, name=name,
            gender=str(rec.get("Gender") or "").strip(),
            birth_date="",                    # age only in xlsx; DOB comes from PDF
            source="seed",
        )
        items: list[EvidenceItem] = []
        summary = str(rec.get("Summary") or "").strip()
        if summary and summary.lower() != "nan":
            items.append(EvidenceItem(evidence_id="E1", kind=EvidenceKind.NOTE,
                                      text=summary, source="staff-entered"))
        seen[pid] = PatientHistory(patient=patient, items=items)
    return list(seen.values())


# ── SOAP PDF reports ──────────────────────────────────────────────────────
def _grab(text: str, label: str, nxt: str) -> str:
    m = re.search(label + r"(.*?)(?=" + nxt + r"|$)", text, re.S)
    return " ".join(m.group(1).split()) if m else ""


def load_pdf_report(path: Path) -> PatientHistory | None:
    path = Path(path)
    if not path.exists():
        return None
    from pypdf import PdfReader

    text = "\n".join(p.extract_text() or "" for p in PdfReader(str(path)).pages)

    def field(pat: str) -> str:
        m = re.search(pat, text)
        return m.group(1).strip() if m else ""

    name = field(r"Patient:?\s*([A-Za-z .]+)")
    if not name:
        return None
    patient = Patient(
        id=_slug(name), name=name,
        gender=field(r"Gender:?\s*(\w+)"),
        birth_date=_dob_to_iso(field(r"DOB:?\s*([0-9/]+)")),
        source="seed",
    )

    subjective = _grab(text, r"Subjective Notes:?", r"Objective Notes:")
    objective = _grab(text, r"Objective Notes:?", r"Assessment Notes:")
    assessment = _grab(text, r"Assessment Notes:?", r"Plan Notes:")
    plan = _grab(text, r"Plan Notes:?", r"Patient MRN|Encounter|$")
    visit_date = field(r"Visit Date:?\s*([0-9/]+)")
    date_iso = _dob_to_iso(visit_date)

    rows = [
        (EvidenceKind.CONDITION, assessment.replace("Diagnosis:", "").strip()),
        (EvidenceKind.NOTE, subjective),
        (EvidenceKind.OBSERVATION, objective),
        (EvidenceKind.NOTE, plan),
    ]
    items = [
        EvidenceItem(evidence_id=f"E{i + 1}", kind=kind, text=txt,
                     date=date_iso, source="staff-entered")
        for i, (kind, txt) in enumerate(rows) if txt
    ]
    return PatientHistory(patient=patient, items=items)


def load_instructor_data() -> list[PatientHistory]:
    """All instructor patients: SOAP PDFs (rich) merged over xlsx roster."""
    by_id: dict[str, PatientHistory] = {h.patient.id: h for h in load_xlsx_patients()}
    for pdf in sorted(DATASETS_DIR.glob("sample_report_*.pdf")):
        hist = load_pdf_report(pdf)
        if hist is not None:                  # PDF is richer -> overrides xlsx row
            by_id[hist.patient.id] = hist
    return list(by_id.values())
